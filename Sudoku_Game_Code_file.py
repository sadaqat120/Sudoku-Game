from tkinter import *
from tkinter import messagebox
from functools import partial
from PIL import ImageTk, Image
import openpyxl as xl
from datetime import datetime

Workbook1 = xl.load_workbook('Scoreboard.xlsx')
WorkSheet2 = Workbook1['Sheet1']

grid = [
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0]
]

solved_grid = [
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0]
]

sudoku_list=[8,5,0 , 0 ,0 ,2 , 4 ,0 ,0,
7, 2 ,0,  0, 0, 0,  0, 0, 9,
0, 0, 4,  0, 0, 0, 0, 0, 0,
0, 0 ,0 ,1 ,0 ,7 , 0, 0, 2
,3, 0, 5,  0, 0, 0 , 9, 0 ,0
,0 ,4 ,0,  0, 0, 0 , 0, 0 ,0,
0, 0, 0 , 0 ,8 ,0 ,0, 7,0,
0,1 ,7 , 0, 0 , 0, 0,0,0 ,
0,0,0,0, 3,6 ,0,4,0]

sudoku_list1=[5,3 ,0 ,0, 7 ,0,0 ,0, 0,
6, 0, 0,1, 9, 5 ,0, 0, 0,
0, 9, 8,  0, 0, 0,  0, 6, 0,
8, 0, 0, 0 ,6 ,0, 0, 0 ,3,
4, 0 ,0, 8, 0, 3, 0, 0, 1,
7 ,0 ,0, 0 ,2 ,0,  0, 0, 6,
0,6 ,0,  0, 0, 0,  2, 8, 0,
0, 0, 0, 4 ,1, 9,  0, 0, 5,
0, 0, 0, 0 ,8, 0, 0 ,7, 9]

sudoku_list2=[
    5, 0, 0, 0, 0, 7, 0, 0, 0,
    0, 0, 0, 1, 0, 0, 0, 3, 0,
    0, 0, 3, 0, 4, 5, 0, 0, 8,
    0, 1, 0, 0, 0, 0, 6, 7, 0,
    0, 0, 7, 0, 0, 0, 0, 0, 0,
    0, 0, 0, 0, 2, 0, 0, 0, 9,
    0, 4, 0, 0, 9, 0, 0, 0, 0,
    0, 0, 0, 5, 0, 0, 7, 0, 0,
    0, 0, 2, 0, 0, 0, 0, 0, 6
]

all_lists = {
    'list1': sudoku_list,
    'list2': sudoku_list1,
    'list3': sudoku_list2
}
all_sudoku_list=list(all_lists.values())

entries = []
solved_sudoku_list=[]
auto_sudoku_list=[]
global_result_achieved=False
sudoku_congrates=0
Name=0
start_time=0
final_time=0

def make_grid(any_list,given_grid):
    index=0
    for i in range(0,9):
        for j in range(0,9):
            given_grid[i][j]=(any_list[index])
            index+=1

def sudoku_solution(given_grid,making_grid):
    if solved_sudoku_list!=[]:
        for s in solved_sudoku_list:
            solved_sudoku_list.remove(s)
    global global_result_achieved
    if global_result_achieved:
        return
    for row in range(0, 9):
        for column in range(0, 9):
            if given_grid[row][column] == 0:
                for number in range(1, 10):
                    if correct_entry(given_grid,row, column, number):
                        given_grid[row][column] = number
                        sudoku_solution(given_grid,making_grid)
                        given_grid[row][column] = 0
                return
    for i in range(0, 9):
        for j in range(0, 9):
            solved_sudoku_list.append(given_grid[i][j])
    make_grid(solved_sudoku_list,making_grid)
    global_result_achieved=True

def correction_for_auto_solver(given_grid,row, column, number):
    count=0
    for i in range(0, 9):
        if given_grid[row][i] == number:
            count+=1
    if count>1:
        return False

    count-=count
    for i in range(0, 9):
        if given_grid[i][column] == number:
            count += 1
    if count>1:
        return False

    count-=count
    a1 = (row // 3) * 3
    a2 = (column // 3) * 3
    for i in range(0, 3):
        for j in range(0, 3):
            if given_grid[a1 + i][a2 + j] == number:
                count+=1
    if count>1:
        return False

    return True

def correct_entry(given_grid,row, column, number):
    for i in range(0, 9):
        if given_grid[row][i] == number:
            return False
    for i in range(0, 9):
        if given_grid[i][column] == number:
            return False
    a1 = (row // 3) * 3
    a2 = (column // 3) * 3
    for i in range(0, 3):
        for j in range(0, 3):
            if given_grid[a1 + i][a2 + j] == number:
                return False
    return True

def grid_print(grid_box):
    for i in range(0, 9):
        for j in range(0, 9):
            print(grid_box[i][j], end=" ")
        print()

def auto_solver_grid():
    create_grid(grid,2)

def check_grid(page):
    global Name
    global sudoku_congrates
    global final_time
    flag = True
    clear_cell_colors()
    for i in range(9):
        for j in range(9):
            value = entries[i][j].get()
            if int(value)==solved_grid[i][j]:
                entries[i][j].config(bg="green")
            else:
                entries[i][j].config(bg="dark red")
                flag=False
    for row in range(9):
        for column in range(9):
            if int(entries[row][column].get())==0:
                flag=False
                break
    if flag:
        current_time1 = datetime.now().time()
        final_time = current_time1.hour * 60 + current_time1.minute
        messagebox.showinfo(Name,"Congartulations !\nYou win the game")
        store_data()
        sudoku_congrates+=1
        page.destroy()
        ask_another_game()

def auto_solver(page):
    auto_sudoku_list = []
    global global_result_achieved
    if solved_sudoku_list!=[]:
        for s in solved_sudoku_list:
            solved_sudoku_list.remove(s)
    for i in range(9):
        for j in range(9):
            value = entries[i][j].get()
            auto_sudoku_list.append(int(value))
    make_grid(auto_sudoku_list,grid)
    for e in range(9):
        for x in range(9):
            if grid[e][x] != 0:
                if correction_for_auto_solver(grid,e,x,grid[e][x]):
                    pass
                else:
                    messagebox.showinfo("Wrong puzzle","Your puzzle is not correct")
                    return
    for z in range(9):
        for y in range(9):
            solved_grid[z][y]=0
    sudoku_solution(grid, solved_grid)
    global_result_achieved = False
    grid_print(solved_grid)
    print()
    for m in range(0,9):
        for n in range(0,9):
            if solved_grid[m][n]==0:
                messagebox.showinfo("No Solution","This puzzle have no solution")
                return
    page.destroy()
    create_grid(solved_grid,3)

def create_puzzle():
    global global_result_achieved
    global sudoku_congrates
    if sudoku_congrates > 0:
        all_sudoku_list.pop(0)
        sudoku_congrates -= 1
    if len(all_sudoku_list)!=0:
        puzzle=all_sudoku_list[0]
        make_grid(puzzle, grid)
        sudoku_solution(grid, solved_grid)
        global_result_achieved = False
        grid_print(solved_grid)
        print()
        New_game(grid)
    else:
        messagebox.showinfo("Well Done !","You have won consecutive games.")

def player_entry():
    entry_page=Tk()
    entry_page.config(bg="white",borderwidth=5)
    entry_page.geometry('800x300')
    entry_page.title("Register Yourself")
    file_path = "sudoku_pic3.png"
    image = Image.open(file_path)
    image = image.resize((810,290))
    photo = ImageTk.PhotoImage(image)
    Label(entry_page, image=photo).place(x=-10,y=70)
    completed = partial(exit_page,entry_page)
    Label(entry_page, text="\t\t\t     Wellcome to Sudoku\t\t\t             ", bg="lime green", font='Helvetica 15 bold',fg='white').place(x=0, y=0)
    Label(entry_page,text="We register player for keeping record and other purposes.\nRegister yourself to play the amazing game of Sudoku!",font='Helvetica 13 bold',bg="white").place(x=170, y=30)
    Label(entry_page, text="Name:", fg='black', bg="white", font="Helvetica 12 bold").place(x=250, y=130)
    enter_name = StringVar()
    Entry(entry_page, textvariable=enter_name, font="Helvitica 12 bold", fg="black",bg="lavender").place(x=310, y=130)
    Button(entry_page, text="Submit", command=partial(entry_completed,enter_name,entry_page), bg="green", bd=10, font="Helvetica 10 bold", fg='white').place(x=350, y=200)
    entry_page.mainloop()

def entry_completed(name, previous_page):
    global Name
    if name.get()=="":
        messagebox.showinfo("Invalid", "Name cannot be None !")
    else:
        Name=name.get()
        messagebox.showinfo(name.get(), "Registered\nThanks !")
        previous_page.destroy()
        sudoku_feature()
    # enter this name anywhere you want ot store for your record ! and make frame to keep record

def sudoku_feature():
    entry_page=Tk()
    entry_page.config(background='grey', borderwidth=5)
    entry_page.geometry('1100x600')
    entry_page.title("Sudoku")
    Label(entry_page, text="\t\t\t\t  SUDOKU\t\t\t\t\t\t\t", bg="lime green", font='Helvetica 20 bold', fg='white').place(x=-10, y=0)
    file_path = "sudoku_pic2.jpg"
    image = Image.open(file_path)
    image = image.resize((1090,550))
    photo = ImageTk.PhotoImage(image)
    # Create a label for the image
    Label(entry_page, image=photo).place(x=0,y=40)
    # Button(entry_page, text="Demo",command=partial(exit_page,entry_page),bg="goldenrod",bd=15,font='Helvetica 18 bold',fg='black').place(x=950, y=60)
    Button(entry_page, text="\tPlay Game\t    ", command=create_puzzle, bg="lime green", font='Helvetica 16 bold',fg='black').place(x=0, y=60)
    Button(entry_page, text="\tInstructions\t    ", command=partial(instruction_box), bg="white", font='Helvetica 16 bold',fg='blue').place(x=0, y=110)
    Button(entry_page, text="\tLeaderboard\t    ", command=leaderboard_call, bg="lime green", font='Helvetica 16 bold',fg='black').place(x=0, y=160)
    Button(entry_page, text="\tSudoku Solver\t    ", command=auto_solver_grid, bg="white", font='Helvetica 16 bold',fg='blue').place(x=0, y=210)
    Button(entry_page, text="Exit", command=partial(exit_page,entry_page), bg="red", bd=8, font='Helvetica 12 bold', fg='white').place(x=1000, y=530)
    entry_page.mainloop()

def instruction_box():
    inst_page=Tk()
    inst_page.config(background='light blue', borderwidth=5)
    inst_page.geometry('600x300')
    inst_page.title("Instruction Box")
    Label(inst_page, text="\t\t           Instructions\t\t\t\t", bg="lime green", font='Helvetica 15 bold', fg='white').place(x=0, y=0)
    Label(inst_page, text="A 9 by 9 grid is used to play Sudoku.There are 9 'squares'(composed\nof 3 x 3 spaces) with in the rows and columns.You must fill out all nine\nslots in each row, column, and square with the numbers 1 to 9,without\nrepeating any number in any row, column, or square.",font='Helvetica 13 bold',bg="light blue").place(x=20, y=50)
    Button(inst_page, text="Okay !", command=partial(destroy_page,inst_page), bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=250, y=170)
    inst_page.mainloop()

def New_game(given_puzzle,):
    global start_time
    current_time2 = datetime.now().time()
    start_time = current_time2.hour * 60 + current_time2.minute
    create_grid(given_puzzle,1)

def create_grid(make_puzzle,signal_check):
    if entries!=[]:
        while len(entries) > 0:
            entries.pop()

    grid_window = Tk()
    grid_window.title("Sudoku")
    grid_window.geometry("560x660")
    for i in range(9):
        row = []
        for j in range(9):
            entry = Entry(grid_window, font=('Arial', 20), bd=2, justify='center', relief="solid",
                             disabledbackground='gray', disabledforeground='black')

            if signal_check==1 or signal_check==3:
                if make_puzzle[i][j] != 0:
                    entry.insert(0, make_puzzle[i][j])
                    entry.config(state='disabled')
                else:
                    entry.insert(0, make_puzzle[i][j])
            elif signal_check==2:
                entry.insert(0,make_puzzle[i][j])
            # Position the entry widget
            entry.place(x=60 * j + 10, y=60 * i + 10, width=60, height=60)
            row.append(entry)
        entries.append(row)
    if signal_check==1 or signal_check==2:
        if signal_check==1:
            Button(grid_window, text="Clear", command=clear_cell_colors, bg="lime green", bd=15,
                                     font='Helvetica 18 bold', fg='white').place(x=10, y=570)
            Button(grid_window, text="Check", command=partial(check_grid,grid_window), bg="lime green", bd=15,
                                     font='Helvetica 18 bold', fg='white').place(x=430, y=570)
        elif signal_check==2:
            Button(grid_window, text="Solve", command=partial(auto_solver,grid_window), bg="lime green", bd=15,
                   font='Helvetica 18 bold', fg='white').place(x=430, y=570)
        Button(grid_window, text="Exit", command=partial(exit_page, grid_window), bg="red", bd=8,
                   font='Helvetica 12 bold', fg='white').place(
                x=250, y=600)
    elif signal_check==3:
        Button(grid_window, text="Okay", command=partial(destroy_page, grid_window), bg="white", bd=15,
               font='Helvetica 12 bold', fg='lime green').place(x=230, y=580)
    grid_window.mainloop()

def ask_another_game():
    another_ganme = Tk()
    another_ganme.config(bg='grey', borderwidth=5)
    another_ganme.geometry('300x200')
    another_ganme.title("Reply")
    Label(another_ganme, text="Play Next Puzzle?", background="white", fg='blue', font="Helvitica 11 bold").place(x=80, y=30)
    Button(another_ganme, text="Yes", command=create_puzzle, bg='green', fg='white', bd=8,
           font='Helvetica 9 bold').place(x=50, y=80)
    Button(another_ganme, text="No", command=partial(exit_page,another_ganme), bg='red', fg='white', bd=8,
           font='Helvetica 9 bold').place(x=200, y=80)
    another_ganme.mainloop()

def clear_cell_colors():
    for i in range(9):
        for j in range(9):
            entries[i][j].config(bg="white")

def exit_page(page):
    messagebox.showinfo("Exit","Game Closed !")
    page.destroy()

def destroy_page(page):
    page.destroy()

def store_data():
    i=1
    while WorkSheet2.cell(row=i, column=1).value != None:
        i += 1
    WorkSheet2.cell(row=i,column=1).value = Name
    WorkSheet2.cell(row=i,column=2).value = final_time-start_time
    Workbook1.save(('Scoreboard.xlsx'))

def leaderboard_call():
    rank=leaderboard()
    show_leaderboard(rank)

def leaderboard():
    data = [(name.value, number.value) for name, number in zip(WorkSheet2['A'], WorkSheet2['B'])]
    filtered_data = [(name, int(number)) for name, number in data if number is not None]
    sorted_data = sorted(filtered_data, key=lambda x: x[1])
    top3 = sorted_data[:3]
    return top3

def show_leaderboard(top3):
    leaderboard = Tk()
    leaderboard.config(bg='lime green', borderwidth=5)
    leaderboard.title("Leaderboard")
    leaderboard.geometry('350x300')
    Label(leaderboard,text="Leaderboard", background="lime green", fg='black', font="Helvitica 18 bold").place(x=80, y=30)
    t=0
    for i, (name, number) in enumerate(top3):
        label_text = f"Rank {i + 1}: \t{name} : {number} minutes"
        Label(leaderboard, text=label_text,background="lime green", fg='black', font="Helvitica 12 bold").place(x=50,y=80+t)
        t+=30
    Button(leaderboard, text="Okay", command=partial(destroy_page, leaderboard), bg="white", bd=11,
           font='Helvetica 12 bold', fg='lime green').place(x=130, y=220)
    leaderboard.mainloop()

player_entry()
